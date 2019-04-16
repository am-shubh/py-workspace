import openpyxl,os
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
print("WELCOME TO EXCEL SHEET HANDLER")
print()
path = raw_input("Enter the path where the excel file is located : ")
os.chdir(path)
print()
print("Directory changed to "+path)
print()
file_name = raw_input("Enter the excel file name : ")
print()
print("Opening "+file_name)
print()
wb = openpyxl.load_workbook(file_name)
sh = wb.get_sheet_names()
print("Sheets present in the file are : ")
for i in sh:
    print(i)
print()
sheet_name = raw_input("Enter the sheet name : ")
print()
sheet = wb.get_sheet_by_name(sheet_name)
print()
print("Focus changed to "+sheet.title)
print()
print("Data available in "+sheet_name+" are :")
print()
for row in sheet.rows:
    for cell in row:
        print(cell.value,)
    print()
print()
print("Choose Desired opotion : ")
print("Press 1 for Getting Cell Value ")
print("Press 2 for Getting Cell Values for some range ")
print("Press 3 for Getting Particular row's and column's Cell Value ")
print()
option = input("Enter Your option : ")

if option==1:
    print()
    r=input("Enter row number : ")
    c=raw_input("Enter column name : ")
    print()
    print("The Cell value in row "+str(r)+" and column "+c+" is :")
    newc=column_index_from_string(c)-1
    print(sheet.cell(row=r-1, column=newc).value)
    print()
elif option==2:
    print()
    start_cell=raw_input("Enter starting cell number : ")
    end_cell=raw_input("Enter end cell number : ")
    print()
    print("All Cell value from "+start_cell+" to "+end_cell+" are : ")
    ir=start_cell[1]
    fr=end_cell[1]
    ic=start_cell[0]
    fc=end_cell[0]
    new_ic=column_index_from_string(ic)-1
    new_fc=column_index_from_string(fc)-1
    r_list=[]
    for i in range(int(ir)-1,int(fr)):
        r_list.append(i)
    for x in r_list:
        counter=new_ic
        for y in sheet.rows[x]:
            if counter>new_fc:
                break
            else:
                print(y.value,)
                counter=counter+1
        print()
    print()
elif option==3:
    print()
    print("Press 1 for getting row values ")
    print("Press 2 for getting column values")
    print()
    opt = input("Enter Your choice : ")
    if opt==1:
        print()
        r=input("Enter Row name : ")
        print()
        print()
        for x in sheet.rows[r-1]:
            print(x.value,)
        print()
    elif opt==2:
        print()
        c=raw_input("Enter Column name : ")
        print()
        print("The cells present in column "+c+" are : ")       
        for x in sheet.columns[column_index_from_string(c)-1]:
            print(x.value)
        print()

print()

















        
