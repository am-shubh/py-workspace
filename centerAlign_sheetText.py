import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = openpyxl.load_workbook("prac.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")
align=Alignment(horizontal='center')
for row in sheet.rows:
    for cell in row:
        cell.alignment=align
wb.save("prac.xlsx")

f=open('test1.txt','r')
for line in f:
    print(int(line))
